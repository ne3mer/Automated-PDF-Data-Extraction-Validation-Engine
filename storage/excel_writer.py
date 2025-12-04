"""
Excel Writer Module
Handles writing extracted data to Excel files
"""
import logging
from pathlib import Path
from typing import List, Dict, Any
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
import config

logger = logging.getLogger(__name__)


class ExcelWriter:
    """Handles writing data to Excel format"""
    
    def __init__(self, output_dir: Path = None):
        """
        Initialize Excel writer
        
        Args:
            output_dir: Output directory for Excel files
        """
        self.output_dir = output_dir or config.EXCEL_OUTPUT_DIR
        self.output_dir.mkdir(parents=True, exist_ok=True)
    
    def write_documents(self, documents: List[Dict[str, Any]], filename: str = "extracted_data.xlsx") -> Path:
        """
        Write documents to Excel file
        
        Args:
            documents: List of document data dictionaries
            filename: Output filename
            
        Returns:
            Path to written Excel file
        """
        if not filename.endswith(".xlsx"):
            filename += ".xlsx"
        
        output_path = self.output_dir / filename
        
        try:
            # Convert to DataFrame
            df = pd.DataFrame(documents)
            
            # Reorder columns for better readability
            preferred_order = [
                "document_id", "document_type", "source_file_name",
                "vendor_name", "client_name", "invoice_number",
                "issue_date", "due_date", "total_amount", "tax_amount",
                "currency", "payment_terms", "reference_number",
                "validation_status", "validation_score", "missing_fields",
                "validation_errors", "processed_timestamp"
            ]
            
            # Reorder columns if they exist
            existing_cols = [col for col in preferred_order if col in df.columns]
            other_cols = [col for col in df.columns if col not in preferred_order]
            df = df[existing_cols + other_cols]
            
            # Write to Excel
            with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
                df.to_excel(writer, sheet_name='Extracted Data', index=False)
                
                # Format the worksheet
                worksheet = writer.sheets['Extracted Data']
                self._format_worksheet(worksheet, df)
            
            logger.info(f"Wrote Excel file with {len(documents)} documents: {output_path}")
            return output_path
            
        except Exception as e:
            logger.error(f"Error writing Excel file {output_path}: {str(e)}")
            raise
    
    def _format_worksheet(self, worksheet, df):
        """Format Excel worksheet with headers and styling"""
        # Header styling
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        
        # Apply header formatting
        for cell in worksheet[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")
        
        # Auto-adjust column widths
        for column in worksheet.columns:
            max_length = 0
            column_letter = column[0].column_letter
            
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            
            adjusted_width = min(max_length + 2, 50)
            worksheet.column_dimensions[column_letter].width = adjusted_width
    
    def write_report(self, report_data: Dict[str, Any], filename: str = "validation_report.xlsx") -> Path:
        """
        Write validation report to Excel
        
        Args:
            report_data: Report data dictionary
            filename: Output filename
            
        Returns:
            Path to written Excel file
        """
        if not filename.endswith(".xlsx"):
            filename += ".xlsx"
        
        output_path = config.REPORTS_OUTPUT_DIR / filename
        
        try:
            # Create summary sheet
            summary_data = {
                "Metric": [
                    "Total Documents Processed",
                    "Successfully Validated (PASSED)",
                    "Partially Validated (PARTIAL)",
                    "Failed Validation (FAILED)",
                    "Documents with Missing Required Fields",
                    "Documents with Date Errors",
                    "Documents with Numeric Errors",
                    "Average Validation Score"
                ],
                "Value": [
                    report_data.get("total_processed", 0),
                    report_data.get("passed", 0),
                    report_data.get("partial", 0),
                    report_data.get("failed", 0),
                    report_data.get("missing_fields_count", 0),
                    report_data.get("date_errors_count", 0),
                    report_data.get("numeric_errors_count", 0),
                    round(report_data.get("average_score", 0.0), 2)
                ]
            }
            
            summary_df = pd.DataFrame(summary_data)
            
            # Create detailed errors sheet if available
            errors_data = report_data.get("error_details", [])
            errors_df = pd.DataFrame(errors_data) if errors_data else pd.DataFrame()
            
            with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
                summary_df.to_excel(writer, sheet_name='Summary', index=False)
                
                if not errors_df.empty:
                    errors_df.to_excel(writer, sheet_name='Error Details', index=False)
                
                # Format summary sheet
                summary_sheet = writer.sheets['Summary']
                self._format_worksheet(summary_sheet, summary_df)
                
                if not errors_df.empty:
                    errors_sheet = writer.sheets['Error Details']
                    self._format_worksheet(errors_sheet, errors_df)
            
            logger.info(f"Wrote validation report Excel: {output_path}")
            return output_path
            
        except Exception as e:
            logger.error(f"Error writing validation report Excel {output_path}: {str(e)}")
            raise

